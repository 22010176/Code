# import string
# from openpyxl import Workbook
# wb = Workbook()

# # grab the active worksheet
# ws = wb.active
# alphabet = [*string.ascii_uppercase]
# CATE = ['funcs', 'aveX', 'aveY', 'SSx', 'SSxy', 'B1',
#         'B0', 'MAE', 'MSE', "R1", 'R2', 'SSR', 'SST']
# funs = ['hello', 'asdfafs', 'asdfasfd', 'asdfasf', 'ase']
# # Data can be assigned directly to cells
# for i in CATE:
#     ws['A'+str(CATE.index(i)+1)] = i.upper()
# # print(alphabet[0])
# for i in funs:
#     ws[f'{alphabet[funs.index(i)+1]}1'] = i.upper()


# # Python types will automatically be converted
# # ws['A2'] = datetime.datetime.now()

# # Save the file
# wb.save("./playground/sample.xlsx")
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()
ws = wb.active

data = [
    ['Apples', 10000, 5000, 8000, 6000],
    ['Pears',   2000, 3000, 4000, 5000],
    ['Bananas', 6000, 6000, 6500, 6000],
    ['Oranges',  500,  300,  200,  700],
]

# add column headings. NB. these must be strings
ws.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    ws.append(row)

tab = Table(displayName="Table1", ref="A1:E5")

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

'''
Table must be added using ws.add_table() method to avoid duplicate names.
Using this method ensures table name is unque through out defined names and all other table name. 
'''
ws.add_table(tab)
wb.save("table.xlsx")
